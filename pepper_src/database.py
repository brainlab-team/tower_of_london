# -*- coding: utf-8 -*-

import sqlite3
from openpyxl import Workbook
from openpyxl import load_workbook

#Funzione che costruisce il database
def createDatabase():
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS TEST(
                testID INTEGER PRIMARY KEY AUTOINCREMENT,
                modality TEXT,
                phase TEXT NOT NULL
            );''')

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS PLAY(
            playId INTEGER PRIMARY KEY AUTOINCREMENT,
            time INTEGER NOT NULL,
            moves INTEGER NOT NULL,
            result TEXT NOT NULL,
            score INT NOT NULL,
            hints INT NOT NULL,
            testID INTEGER,
            userID INTEGER NOT NULL,
            minimumMoves INTERGER NOT NULL,
            FOREIGN KEY(testID) REFERENCES TEST(testID),
            UNIQUE(playId, userID)
        );''')

        #Definizione delle varie fasi del test
        test1 = """INSERT INTO TEST (modality, phase) VALUES ('','Baseline');"""
        test2 = """INSERT INTO TEST (modality, phase) VALUES ('Full support','Training');"""
        test3 = """INSERT INTO TEST (modality, phase) VALUES ('Fostering Autonomy','Training');"""
        test4 = """INSERT INTO TEST (modality, phase) VALUES ('','Assessment');"""
        cursor.execute(test1)
        cursor.execute(test2)
        cursor.execute(test3)
        cursor.execute(test4)
        connection.commit()
        connection.close()
    except Exception as e:
        print("Error during the database creation [%s]" % e)

#Funzione che si occupa di inserire un nuovo record di PLAY
def insertPlay(time, moves, result, score, hints, testID, userID, minimumMoves):
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        cursor.execute("INSERT INTO PLAY (time, moves, result, score, hints, testID, userID, minimumMoves) VALUES (?, ?, ?, ?, ?, ?, ?, ?)", (time, moves, result, score, hints, testID, userID, minimumMoves))
        connection.commit()
        connection.close()
    except Exception as e:
        print("Error inserting play into the database.")
        print(e)

#Funzione che stampa i test presente nel db
def printTests():
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM TEST")
        tests = cursor.fetchall()
        connection.close()

        if tests:
            print("Tests:")
            for test in tests:
                print("Test ID: {test[0]}, Modality: {test[1]}, Phase: {test[2]}".format(test=test))  # Modifica qui
        else:
            print("No tests found in the database.")
    except Exception as e:
        print("Error retrieving tests from the database.")
        print(e)

#Funzione che stampa le partite presente nel db
def printPlays():
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM PLAY")
        plays = cursor.fetchall()
        connection.close()

        if plays:
            print("Plays:")
            for play in plays:
                print("Play ID: {play[0]}, Time: {play[1]}, Moves: {play[2]}, Result: {play[3]}, Score: {play[4]}, Hints: {play[5]}, Test ID: {play[6]}, User ID: {play[7]}".format(play=play))
        else:
            print("No plays found in the database.")
    except Exception as e:
        print("Error retrieving plays from the database.")
        print(e)

#Funzione che salva i dati presenti sul db nel file risultati.xlsx 
def printGames():
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        cursor.execute("SELECT PLAY.userID, TEST.modality, TEST.phase, PLAY.time, PLAY.result, PLAY.minimumMoves, PLAY.moves, PLAY.hints, PLAY.score FROM PLAY JOIN TEST ON PLAY.testID = TEST.testID;")
        games = cursor.fetchall()
        
        if games:
            # Carica il foglio di lavoro Excel esistente se presente
            try:
                wb = load_workbook("results.xlsx")
                ws = wb.active
            except:
                # Se il file non esiste, crea un nuovo foglio di lavoro
                wb = Workbook()
                ws = wb.active
                # Aggiungi intestazioni solo se il foglio di lavoro è nuovo
                ws.append(['UserID', 'Modalità', 'Fase','Tempo', 'Risultato', 'Mosse necessarie','Mosse effettuate', 'Suggerimenti', 'Punteggio'])

            first_empty_row = 1
            while ws.cell(row=first_empty_row, column=1).value is not None:
                first_empty_row += 1

            for col, value in enumerate(games[-1], start=1):
                ws.cell(row=first_empty_row, column=col).value = value

            # Salvataggio del foglio di lavoro in un file Excel
            wb.save("results.xlsx")
            print("Data saved on file")
        else:
            print("No games found in the database.")
    except Exception as e:
        print("Error retrieving games from the database.")
        print(e)

#Funzione per modificare un valore in una colonna di una tabella del db
def updateValueFromTable(nameOfTable, column, oldValue, newValue):
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        # Utilizzo del parametro per il nome della tabella e del campo da aggiornare
        cursor.execute("UPDATE {table} SET {column} = ? WHERE {column} = ?".format(table=nameOfTable, column=column), (newValue, oldValue))
        
        # Esecuzione della query e salvataggio dei risultati
        connection.commit()
        print("Value updated successfully.")
    except Exception as e:
        print("Error updating value in the database.")
        print(e)
    finally:
        # Chiusura della connessione al database
        connection.close()

#Funzione per eliminare un record di una tabella del db
def deleteValueFromTable(tableName, columnName, value):
    try:
        connection = sqlite3.connect("pepperDatabase.db")
        cursor = connection.cursor()
        
        # Utilizzo dei parametri per il nome della tabella, della colonna e del valore da eliminare
        cursor.execute("DELETE FROM {table} WHERE {column} = ?;".format(table=tableName, column=columnName), (value,))
        
        # Esecuzione della query e salvataggio dei risultati
        connection.commit()
        print("Value deleted successfully.")
    except Exception as e:
        print("Error deleting value from the database.")
        print(e)
    finally:
        # Chiusura della connessione al database
        connection.close()



